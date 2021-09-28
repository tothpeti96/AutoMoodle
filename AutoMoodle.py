import openpyxl
import sys
import tkinter as tk
from tkinter import filedialog
from tqdm import tqdm

root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename()

def WorkBook(wb_name):
    wb = openpyxl.load_workbook(wb_name, data_only = True)
    return wb

class Sheet():
    def __init__(self, wb, sheet_name):
        self.wb = wb
        self.sheet_name = sheet_name
        self.sheet = None
        self.text_length = None
        self.matrix_dim = None
        self.Questions = []
        self.Data = []
        self.variables = []
        self.equation = None

    def ReadSheet(self):
        self.sheet = self.wb[self.sheet_name]

    def TextDim(self):
        sheet = self.sheet
        row_num = 0
        i = 2
        while True:
            if sheet.cell(row = i, column = 1).value is not None:
                row_num += 1
            else:
                break
            i+=1
        self.text_length = row_num

    def ParamDim(self):

        row_num = 0
        i = 4 + self.text_length
        while True:
            if self.sheet.cell(row = i, column = 1).value is not None:
                row_num += 1
            else:
                break
            i += 1

        col_num = 0

        i = 2
        while True:
            if self.sheet.cell(row = 4 + self.text_length, column = i).value is not None:
                col_num += 1
            else:
                break
            i += 1

        self.matrix_dim = (row_num, col_num)

    def ReadData(self):
        length = self.text_length / 10
        for i in range(int(length)):
            title = self.sheet.cell(row = 2 + i * 10, column = 2).value
            question = self.sheet.cell(row = 3 + i * 10, column = 2).value
            help = self.sheet.cell(row = 4 + i * 10, column = 2).value
            unit = self.sheet.cell(row = 5 + i * 10, column = 2).value
            answer = self.sheet.cell(row = 6 + i * 10, column = 2).value
            point = self.sheet.cell(row = 7 + i * 10, column = 2).value
            tolerance = self.sheet.cell(row = 8 + i * 10, column = 2).value

            if self.sheet.cell(row = 9 + i * 10, column = 2).value == 'r':
                tT = 1
            elif self.sheet.cell(row = 9 + i * 10, column = 2).value == 'n':
                tT = 2
            elif self.sheet.cell(row = 9 + i * 10, column = 2).value == 'm':
                tT = 3
            toleranceType = tT

            if self.sheet.cell(row = 10 + i * 10, column = 2).value == 's':
                cF = 2
            elif self.sheet.cell(row = 10 + i * 10, column = 2).value == 'd':
                cF = 1
            format = cF
                
            decimal = self.sheet.cell(row = 11 + i * 10, column = 2).value
            
            self.Questions.append((title, question, help, unit, answer, 
                                   point, tolerance, toleranceType, format, decimal))

        rows = self.matrix_dim[0]
        columns = self.matrix_dim[1]
        
        for i in range(rows):
            var = self.sheet.cell(row = 4 + self.text_length + i, column = 1).value
            self.variables.append(var)
        
        for i in range(columns):
            dataset = {}
            for j in range(rows):
                dataset[self.variables[j]] = self.sheet.cell(row = 4 + self.text_length + j, column = 2 + i).value
            self.Data.append(dataset)


    def SetEquation(self):
        eq_par = []
        for i in range(len(self.variables)):
            eq_par.append('{}{}{}'.format('{',self.variables[i],'}'))
        str_eq = '*'.join(eq_par)
        self.equation = str_eq


    def SetData(self, file):
        for i in range(len(self.variables)):
            file.write("""<dataset_definition>
                <status><text>shared</text>
                </status>
                <name><text>{}</text>
                </name>
                <type>calculated</type>
                <distribution><text>uniform</text>
                </distribution>
                <minimum><text>0</text>
                </minimum>
                <maximum><text>10</text>
                </maximum>
                <decimals><text>4</text>
                </decimals>
                <itemcount>{}</itemcount>
                <dataset_items>\n""".format(self.variables[i], len(self.Data)))

            for j in range(len(self.Data)):
                file.write("""<dataset_item>
                            <number>{}</number>
                            <value>{}</value>
                            </dataset_item>\n""".format(j+1, round(self.Data[j][self.variables[i]],4)))

            
            file.write("""</dataset_items>
                            <number_of_items>{}</number_of_items>
                            </dataset_definition>\n""".format(len(self.Data)))
            
class Test():
    def __init__(self, file, exc, variable, equation):
        self.file = file 
        self.exc = exc  
        self.variable = variable
        self.equation = equation
        
    def TestName(self):
        self.file.write("""<question type="calculated">
                    <name>
                    <text>{}</text>
                    </name>\n""".format(self.exc[0]))

    def TestQuestion(self):
        self.file.write("""<questiontext format="html">
                        <text><![CDATA[<p dir="ltr" style="text-align: left;"></p><p></p><p>{}<p></p><br><p></p>]]></text>
                        </questiontext>""".format(self.exc[1]))

    def FeedBack(self):
        self.file.write("""<generalfeedback format="html">
                    <text><![CDATA[Megoldásmenet:&nbsp;<br>{}&nbsp;<p dir="ltr" style="text-align: left;"></p><p dir="ltr"></p><p></p>]]></text>
                    </generalfeedback>""".format(self.exc[2]))

    def Grade(self):
        self.file.write("""<defaultgrade>{}</defaultgrade>
                <penalty>0.3333333</penalty>
                <hidden>0</hidden>
                <idnumber></idnumber>
                <synchronize>1</synchronize>
                <single>0</single>
                <answernumbering>abc</answernumbering>
                <shuffleanswers>1</shuffleanswers>
                <correctfeedback>
                <text></text>
                </correctfeedback>
                <partiallycorrectfeedback>
                <text></text>
                </partiallycorrectfeedback>
                <incorrectfeedback>
                <text></text>
                </incorrectfeedback>\n""".format(self.exc[5]))

    def Answer(self):
        self.file.write("""<answer fraction="100">
                    <text>{}{}{}</text>
                    <tolerance>{}</tolerance>
                    <tolerancetype>{}</tolerancetype>
                    <correctanswerformat>{}</correctanswerformat>
                    <correctanswerlength>{}</correctanswerlength>
                    <feedback format="html">
                    <text></text>
                    </feedback>
                    </answer>\n""".format('{',self.variable,'}', self.exc[6], self.exc[7], self.exc[8], self.exc[9]))


    def Equation(self):
        self.file.write("""<answer fraction="0">
            <text>{}</text>
            <tolerance>1</tolerance>
            <tolerancetype>1</tolerancetype>
            <correctanswerformat>1</correctanswerformat>
            <correctanswerlength>2</correctanswerlength>
            <feedback format="html">
            <text></text>
            </feedback>
            </answer>\n""".format(self.equation))

    def Penalty(self):
        self.file.write("""<unitgradingtype>1</unitgradingtype>
            <unitpenalty>1.0000000</unitpenalty>
            <showunits>3</showunits>
            <unitsleft>0</unitsleft>\n""")
        

    def StartDataSet(self):
        self.file.write('<dataset_definitions>\n')


    def FinishDataSet(self):
        self.file.write('</dataset_definitions>\n')

    def SetUnits(self):
        
        if self.exc[3] is None:
            self.file.write("""<unitgradingtype>0</unitgradingtype>
                            <unitpenalty>0.1000000</unitpenalty>
                            <showunits>3</showunits>
                            <unitsleft>0</unitsleft>""")


        
        else:
            self.file.write("""<unitgradingtype>1</unitgradingtype>
                            <unitpenalty>1.0000000</unitpenalty>
                            <showunits>2</showunits>
                            <unitsleft>0</unitsleft>""")

            
            self.file.write('<units>')
            UNITS = []
            units = self.exc[3]
            split1 = units.split()
            for i in range(len(split1)):
                split2 = split1[i].split(',')
                UNITS.append((split2[0],split2[1]))

            for i in range(len(UNITS)):
                self.file.write("""<unit>
                <multiplier>{}</multiplier>
                <unit_name>{}</unit_name>
                </unit>\n""".format(UNITS[i][1], UNITS[i][0]))
            
            self.file.write('</units>')
        

    def EndQuestion(self):
            self.file.write('</question>')

def StartFile(FileName):
    f = open(FileName, 'wt', encoding = 'utf-8')
    f.write('<?xml version="1.0" encoding="UTF-8"?>\n<quiz>\n')
    return f
    
def FinishFile(file):
    file.write('</quiz>')
    file.close()

def main(wb, sheet_name):

    FileName = sheet_name + '_moodle.xml'
    sheet = Sheet(wb, sheet_name)
    sheet.ReadSheet()
    sheet.TextDim()
    sheet.ParamDim()
    sheet.ReadData()
    sheet.SetEquation()

    file = StartFile(FileName)
    
    print("Generating {}:".format(FileName))

    for i in tqdm(range(len(sheet.Questions))):
        test = Test(file, sheet.Questions[i], sheet.Questions[i][4], sheet.equation)
        test.TestName()
        test.TestQuestion()
        test.FeedBack()
        test.Grade()
        test.Answer()
        test.Equation()
        test.SetUnits()
        test.StartDataSet()
        sheet.SetData(file)
        test.FinishDataSet()
        test.EndQuestion()

    FinishFile(file)

    print()


print()

wb = WorkBook(file_path)
sheet_names = wb.sheetnames
for sheet in sheet_names:
    main(wb, sheet)